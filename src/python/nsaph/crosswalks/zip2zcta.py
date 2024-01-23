#  Copyright (c) 2023. Harvard University
#
#  Developed by Research Software Engineering,
#  Faculty of Arts and Sciences, Research Computing (FAS RC)
#  Author: Michael A Bouzinier
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#         http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
Utility to download crosswalk files from https://udsmapper.org/zip-code-to-zcta-crosswalk/

Alternative source could be:
[Census Reporter](https://github.com/censusreporter/acs-aggregate/blob/master/crosswalks/zip_to_zcta/ZIP_ZCTA_README.md)
"""
import gzip
import json
import logging
import os.path
import sys
from argparse import ArgumentParser
from typing import Dict, List

import openpyxl
import xlrd

from nsaph.util.pg_json_dump import populate
from nsaph.util.resources import get_resources
from nsaph_utils.utils.io_utils import download, HEADERS
from openpyxl.worksheet.worksheet import Worksheet

from nsaph import init_logging
from nsaph.db import Connection
import requests


class XReader:
    COLUMNS = ["ZIP_CODE", "PO_NAME", "STATE", "ZIP_TYPE", "ZCTA", "zip_join_type"]
    ignored_columns = {"StateName", "OBJECTID", "ENC_ZIP", "ReportingYear"}
    column_map = {
        "ZIP": "ZIP_CODE",
        "ZIPType": "ZIP_TYPE",
        "CityName": "PO_NAME",
        "StateAbbr": "STATE",
        "ZCTA_USE": "ZCTA",
        "Zip_join_type": "zip_join_type"
    }

    
    def __init__(self, path_to_file: str):
        self.path: str = path_to_file
        self.year = None
        return

    def get_year(self):
        for y in range(1990,2100):
            if str(y) in self.path:
                self.year = y
                break
        return

    @staticmethod
    def ensure_join_type(columns: List[str], values: Dict[str,str]):
        if "zip_join_type" not in columns:
            if values["ZIP_CODE"] == values["ZCTA"]:
                values["zip_join_type"] = "Zip matches ZCTA"
            else:
                values["zip_join_type"] = "Spatial join to ZCTA"

    def read(self) -> List[Dict]:
        if self.year is None:
            self.get_year()
        if self.year is None:
            raise ValueError("Unknown year for file: " + self.path)
        if self.path.endswith(".xls"):
            return self.read_xls()
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws: Worksheet = None
        for s in wb.sheetnames:
            sheet: Worksheet = wb[s]
            if sheet.max_row > 1000:
                ws = sheet
                break
        if ws is None:
            raise ValueError("Crosswalk table is not found inside workbook: " + self.path)
        column_range = range(ws.min_column, ws.max_column + 1)
        titles = [
            ws.cell(ws.min_row, i).value for i in column_range
        ]
        columns = self.map(titles)
        logging.info("Processing: " + self.path)
        data: List[Dict] = []
        for row in ws.iter_rows(min_row=ws.min_row + 1):
            values = {
                columns[i - 1]: row[i - 1].value for i in column_range
                    if columns[i - 1] not in self.ignored_columns
            }
            values["year"] = self.year
            self.ensure_join_type(columns, values)
            data.append(values)
        return data

    def map(self, columns: List[str]) -> List[str]:
        mapped = []
        for c in columns:
            if c in self.COLUMNS:
                mapped.append(c)
            elif c in self.column_map:
                mapped.append(self.column_map[c])
            elif c in self.ignored_columns:
                mapped.append(c)
                continue
            else:
                raise ValueError("Unknown column: " + c)
        return mapped



    def read_xls(self) -> List[Dict]:
        if self.year is None:
            self.get_year()
        if self.year is None:
            raise ValueError("Unknown year for file: " + self.path)
        wb = xlrd.open_workbook(self.path)
        ws = None
        for s in wb.sheet_names():
            sheet = wb.sheet_by_name(s)
            if sheet.nrows > 1000:
                ws = sheet
                break
        if ws is None:
            raise ValueError("Crosswalk table is not found inside workbook: " + self.path)
        column_range = range(0, ws.ncols)
        titles = [
            ws.cell_value(0, i) for i in column_range
        ]
        columns = self.map(titles)
        logging.info("Processing: " + self.path)
        data: List[Dict] = []
        for rowx in range(1, ws.nrows):
            values = {
                columns[colx]: ws.cell_value(rowx, colx) for colx in column_range
            }
            values["year"] = self.year
            self.ensure_join_type(columns, values)
            data.append(values)
        return data


class Importer:
    URL_PATTERN = "https://udsmapper.org/wp-content/uploads/2022/10/Z{ip}CodetoZCTACrosswalk{year}UDS.xls{x}"
    YEARS = [y for y in range(2009, 2022)]
    TABLE = "public.zip2zcta"

    def __init__(self, arguments):
        self.db = arguments.db
        self.conn = arguments.connection
        return

    def download(self, year) -> str:
        url = None
        for spelling in ["ip", "IP"]:
            for ext in ["", "x"]:
                u = self.URL_PATTERN.format(year = year, ip = spelling, x = ext)
                logging.debug("Trying: " + u)
                response = requests.get(u, headers=HEADERS)
                if response.ok:
                    url = u
                    break
        if url is None:
            raise ValueError("Cannot find URL for year " + str(year))
        xlsx = os.path.basename(url)
        if not os.path.isfile(xlsx):
            logging.info("Downloading: " + url)
            with open(xlsx, "wb") as out:
                download(url, out)
        else:
            logging.info("Using existing file: " + xlsx)
        return xlsx

    def process(self, output):
        with gzip.open(output, "wt") as out:
            for y in self.YEARS:
                xlsx = self.download(y)
                reader = XReader(xlsx)
                data = reader.read()
                for line in data:
                    json.dump(line, out)
                    out.write("\n")
        return

    def drop(self):
        self.execute_db_update("DROP TABLE IF EXISTS {} CASCADE".format(self.TABLE))

    def create(self):
        resource = get_resources(self.TABLE)
        ddl_path = resource['ddl']
        with open(ddl_path) as f:
            ddl = ''.join([
                line for line in f
            ])
        self.execute_db_update(ddl)
        return

    def populate(self, path_to_data: str):
        with Connection(self.db, self.conn) as cnxn:
            with cnxn.cursor() as cursor:
                populate(cursor, self.TABLE, path_to_data)
            cnxn.commit()
        #self.execute_db_update()

    def vacuum(self):
        self.execute_db_update("VACUUM (VERBOSE, PARALLEL 6, ANALYZE) public.zip2zcta;")

    def ingest(self, path_to_data: str):
        self.drop()
        self.create()
        self.populate(path_to_data)
        self.vacuum()
        return

    def execute_db_update(self, sql: str):
        with Connection(self.db, self.conn) as cnxn:
            with cnxn.cursor() as cursor:
                logging.info(sql)
                cursor.execute(sql)
            cnxn.commit()

if __name__ == '__main__':
    os.system("rm *.log")
    init_logging()
    parser = ArgumentParser (description="Import ZIP code to ZCTA mappings (crosswalk)")
    parser.add_argument("--db",
                        help="Path to a database connection parameters file",
                        default="database.ini",
                        required=False)
    parser.add_argument("--connection", "-c",
                        help="Section in the database connection parameters file",
                        default="nsaph2",
                        required=False)
    parser.add_argument("--action", "-a",
                        help="",
                        choices=["download", "ingest", "all"],
                        default="all",
                        required=False)

    arguments = parser.parse_args()

    importer = Importer(arguments)
    if arguments.action in ["download", "all"]:
        importer.process("zip2zcta.json.gz")
    if arguments.action in ["ingest", "all"]:
        importer.ingest("zip2zcta.json.gz")

    